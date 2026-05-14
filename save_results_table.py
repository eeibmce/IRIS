# save_results_table.py
# Two jobs:
#   1. print_terminal_result()  - prints a full detailed breakdown to the terminal
#   2. save_results_table()     - writes the result into Results_Table.xlsx,
#                                 creating the file from scratch if it does not exist

import os
import config


# ── Terminal printing ─────────────────────────────────────────────────────────

def print_terminal_result(result):
    # Prints a full breakdown of every field to the terminal.
    # Called automatically after every analysis run.

    gaze_score    = result.get("gaze_score",          0.0)
    nlp_score     = result.get("nlp_score",           0.0)
    gaze_w        = result.get("fusion_gaze_weight",   config.GAZE_WEIGHT)
    nlp_w         = result.get("fusion_nlp_weight",    config.NLP_WEIGHT)
    weighted_sum  = result.get("weighted_sum",         result.get("raw_score", 0.0))
    perceptron    = result.get("perceptron_output",    1 if weighted_sum >= 0.5 else 0)
    decision      = result.get("decision",             "unknown")
    confidence    = result.get("confidence",           0.0)
    uncertainty   = result.get("uncertainty",          0.0)
    rule_name     = result.get("rule_engine_rule",     "unknown")
    prompt        = result.get("closed_loop_iterations", 0) > 0
    iterations    = result.get("closed_loop_iterations", 0)

    nlp_features     = result.get("nlp_features", {})
    filler_rate      = nlp_features.get("filler_rate",       0.0)
    repetition       = nlp_features.get("repetition",        0.0)
    length_variation = nlp_features.get("length_variation",  0.0)
    nlp_anomaly      = nlp_features.get("nlp_anomaly_score", nlp_score)

    gaze_features    = result.get("gaze_features", {})
    eye_contact      = gaze_features.get("eye_contact_ratio",  0.0)
    gaze_consistency = gaze_features.get("gaze_consistency",   0.0)
    blink_rate       = gaze_features.get("blink_rate_per_min", 0.0)
    kp_conf          = gaze_features.get("avg_kp_confidence",  0.0)
    valid_frames     = gaze_features.get("valid_frame_count",  0)
    total_frames     = gaze_features.get("frame_count",        0)

    w1 = config.NLP_FILLER_WEIGHT
    w2 = config.NLP_REPETITION_WEIGHT
    w3 = config.NLP_VARIATION_WEIGHT

    if decision == "uncertain":
        system_decision = "Uncertain - Awaiting Prompt" if not prompt else "Uncertain - Prompted"
    else:
        system_decision = "Anomaly" if perceptron == 1 else "Normal"
    prompt_text     = "Yes" if prompt else "No"
    final_map = {
        "real":      "Genuine",
        "fake":      "Anomaly Found",
        "uncertain": "Uncertain - Prompt Triggered" if prompt else "Uncertain",
    }
    final_classification = final_map.get(decision, decision.title())

    W = 62   # total width of the box

    def line(label, value, indent=2):
        # Print one label: value line inside the box
        pad = " " * indent
        print("|" + pad + str(label).ljust(32) + str(value).ljust(W - 34 - indent) + "|")

    def divider(title=""):
        if title:
            side = (W - len(title) - 2) // 2
            print("+" + "-" * side + " " + title + " " + "-" * (W - side - len(title) - 2) + "+")
        else:
            print("+" + "-" * W + "+")

    def blank():
        print("|" + " " * W + "|")

    divider()
    title_text = "INTERVIEW ANALYSIS RESULT"
    print("|" + title_text.center(W) + "|")
    divider()

    # ── Final Decision ────────────────────────────────────────────────────────
    blank()
    line("Final Classification:", final_classification)
    line("System Decision:",      system_decision + "  (perceptron = " + str(perceptron) + ")")
    line("Decision:",             decision.upper())
    line("Confidence:",           str(round(confidence * 100, 1)) + "%")
    line("Uncertainty:",          str(round(uncertainty * 100, 1)) + "%")
    line("Prompt Triggered:",     prompt_text + "  (iterations: " + str(iterations) + ")")

    # Show clear anomaly notes if present
    notes = result.get("clear_anomaly_notes", [])
    if notes:
        blank()
        print("|  " + "Clear anomaly findings (prompt not triggered):".ljust(W - 2) + "|")
        for note in notes:
            txt = "  - " + note
            print("|" + txt.ljust(W) + "|")
    blank()

    # ── Fusion / Rule Engine ──────────────────────────────────────────────────
    divider("Fusion  /  Rule Engine")
    blank()
    line("Rule fired:",           rule_name)
    line("Fusion weights (w1,w2):", "(" + str(gaze_w) + ",  " + str(nlp_w) + ")")
    line("Eye-Gaze Score (x1):",  str(round(gaze_score, 4)))
    line("NLP Anomaly Score (x2):", str(round(nlp_score, 4)))
    line("Weighted Sum (w1*x1 + w2*x2):", str(round(weighted_sum, 4)))
    blank()

    # ── Gaze Features ─────────────────────────────────────────────────────────
    divider("Gaze  Features")
    blank()
    line("Eye Contact Ratio:",    str(round(eye_contact * 100, 1)) + "%")
    line("Gaze Consistency:",     str(round(gaze_consistency, 3)))
    line("Blink Rate:",           str(round(blink_rate, 1)) + " blinks/min")
    line("YOLO KP Confidence:",   str(round(kp_conf, 3)))
    line("Valid Frames:",         str(valid_frames) + " / " + str(total_frames))
    blank()

    # ── NLP Features ──────────────────────────────────────────────────────────
    divider("NLP  Features")
    blank()
    line("NLP Weights (w1,w2,w3):", "(" + str(w1) + ",  " + str(w2) + ",  " + str(w3) + ")")
    line("Filler Rate (x1):",     str(round(filler_rate, 4)) +
         "  [" + _band_label(filler_rate, 0.05, 0.10) + "]")
    line("Repetition (x2):",      str(round(repetition, 4)) +
         "  [" + _band_label(repetition, 0.10, 0.20) + "]")
    line("Length Variation (x3):", str(round(length_variation, 4)) +
         "  [" + _band_label(length_variation, 0.30, 0.60) + "]")
    line("NLP Weighted Sum:",     str(round(w1*filler_rate + w2*repetition + w3*length_variation, 4)))
    line("NLP Anomaly Score:",    str(round(nlp_anomaly, 4)))
    blank()

    # ── Evidence ──────────────────────────────────────────────────────────────
    divider("Evidence")
    blank()
    evidence = sorted(result.get("evidence", []), key=lambda e: e["weight"], reverse=True)
    if evidence:
        for item in evidence:
            desc = item["description"]
            wt   = "w=" + str(round(item["weight"], 2))
            # Truncate long descriptions to fit in box
            max_desc = W - 10
            if len(desc) > max_desc:
                desc = desc[:max_desc - 3] + "..."
            print("|  [" + item["type"][:4] + "] " + desc.ljust(max_desc) + " " + wt + "|")
    else:
        line("No evidence items recorded.", "")
    blank()

    # ── Transcript preview ────────────────────────────────────────────────────
    transcript = result.get("transcript", "")
    if transcript:
        divider("Transcript  Preview")
        blank()
        preview = transcript[:300]
        if len(transcript) > 300:
            preview += "..."
        # Word-wrap to fit inside box
        words = preview.split()
        current_line = ""
        for word in words:
            if len(current_line) + len(word) + 1 > W - 4:
                print("|  " + current_line.ljust(W - 3) + "|")
                current_line = word
            else:
                current_line = (current_line + " " + word).strip()
        if current_line:
            print("|  " + current_line.ljust(W - 3) + "|")
        blank()

    divider()
    print("")


# ── Spreadsheet ───────────────────────────────────────────────────────────────

def save_results_table(result, xlsx_path="Results_Table.xlsx", case_id=None):
    # Writes the result into the spreadsheet.
    # Creates the file from scratch if it does not exist.

    try:
        from openpyxl import load_workbook, Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl")
        return

    # Create the file if it does not exist
    if not os.path.exists(xlsx_path):
        print("Spreadsheet not found - creating: " + xlsx_path)
        _create_blank_spreadsheet(xlsx_path)

    wb = load_workbook(xlsx_path)

    # Make sure both sheets exist (handles partial files)
    if "Results Table" not in wb.sheetnames:
        _add_results_sheet(wb)
    if "NLP" not in wb.sheetnames:
        _add_nlp_sheet(wb)

    # ── Extract values ────────────────────────────────────────────────────────

    gaze_score    = result.get("gaze_score",          0.0)
    nlp_score     = result.get("nlp_score",           0.0)
    gaze_w        = result.get("fusion_gaze_weight",   config.GAZE_WEIGHT)
    nlp_w         = result.get("fusion_nlp_weight",    config.NLP_WEIGHT)
    weighted_sum  = result.get("weighted_sum",         result.get("raw_score", 0.0))
    perceptron    = result.get("perceptron_output",    1 if weighted_sum >= 0.5 else 0)
    decision      = result.get("decision",             "unknown")
    prompt        = result.get("closed_loop_iterations", 0) > 0

    if decision == "uncertain":
        system_decision = "Uncertain - Awaiting Prompt" if not prompt else "Uncertain - Prompted"
    else:
        system_decision = "Anomaly" if perceptron == 1 else "Normal"
    prompt_text          = "Yes" if prompt else "No"
    final_map = {
        "real":      "Genuine",
        "fake":      "Anomaly Found",
        "uncertain": "Uncertain - Prompt Triggered" if prompt else "Uncertain",
    }
    final_classification = final_map.get(decision, decision.title())
    weights_str          = "(" + str(gaze_w) + ", " + str(nlp_w) + ")"

    nlp_features     = result.get("nlp_features", {})
    filler_rate      = nlp_features.get("filler_rate",       0.0)
    repetition       = nlp_features.get("repetition",        0.0)
    length_variation = nlp_features.get("length_variation",  0.0)
    nlp_anomaly      = nlp_features.get("nlp_anomaly_score", nlp_score)

    w1 = config.NLP_FILLER_WEIGHT
    w2 = config.NLP_REPETITION_WEIGHT
    w3 = config.NLP_VARIATION_WEIGHT
    nlp_weights_str  = "(" + str(w1) + ", " + str(w2) + ", " + str(w3) + ")"
    nlp_weighted_sum = round(w1 * filler_rate + w2 * repetition + w3 * length_variation, 4)

    # ── Results Table sheet ───────────────────────────────────────────────────

    ws_results    = wb["Results Table"]
    DATA_START    = 5
    DATA_END      = 9

    target_row = None
    if case_id is not None:
        target_row = DATA_START + (case_id - 1)
    else:
        for row in range(DATA_START, DATA_END + 1):
            if ws_results.cell(row=row, column=5).value is None:
                target_row = row
                break
    if target_row is None:
        # All rows full - append a new one
        target_row = DATA_END + 1

    case_number = target_row - DATA_START + 1

    ws_results.cell(row=target_row, column=4).value  = case_number
    ws_results.cell(row=target_row, column=5).value  = round(gaze_score, 4)
    ws_results.cell(row=target_row, column=6).value  = round(nlp_score, 4)
    ws_results.cell(row=target_row, column=7).value  = weights_str
    ws_results.cell(row=target_row, column=8).value  = round(weighted_sum, 4)
    ws_results.cell(row=target_row, column=9).value  = perceptron
    ws_results.cell(row=target_row, column=10).value = system_decision
    ws_results.cell(row=target_row, column=11).value = prompt_text
    ws_results.cell(row=target_row, column=12).value = final_classification

    fill_colour = "FFD7D7" if perceptron == 1 else "D7FFD7"
    fill = PatternFill("solid", start_color=fill_colour, end_color=fill_colour)
    for col in range(4, 13):
        cell = ws_results.cell(row=target_row, column=col)
        cell.fill      = fill
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── NLP sheet ─────────────────────────────────────────────────────────────

    ws_nlp       = wb["NLP"]
    NLP_START    = 45
    NLP_END      = 49

    nlp_row = None
    if case_id is not None:
        nlp_row = NLP_START + (case_id - 1)
    else:
        for row in range(NLP_START, NLP_END + 1):
            if ws_nlp.cell(row=row, column=6).value is None:
                nlp_row = row
                break
    if nlp_row is None:
        nlp_row = NLP_END + 1

    ws_nlp.cell(row=nlp_row, column=5).value  = nlp_row - NLP_START + 1
    ws_nlp.cell(row=nlp_row, column=6).value  = round(filler_rate, 4)
    ws_nlp.cell(row=nlp_row, column=7).value  = _band_label(filler_rate, 0.05, 0.10)
    ws_nlp.cell(row=nlp_row, column=8).value  = round(repetition, 4)
    ws_nlp.cell(row=nlp_row, column=9).value  = _band_label(repetition, 0.10, 0.20)
    ws_nlp.cell(row=nlp_row, column=10).value = round(length_variation, 4)
    ws_nlp.cell(row=nlp_row, column=11).value = _band_label(length_variation, 0.30, 0.60)
    ws_nlp.cell(row=nlp_row, column=12).value = nlp_weights_str
    ws_nlp.cell(row=nlp_row, column=13).value = nlp_weighted_sum

    nlp_fill = PatternFill("solid", start_color="EAF3FB", end_color="EAF3FB")
    for col in range(5, 14):
        cell = ws_nlp.cell(row=nlp_row, column=col)
        cell.fill      = nlp_fill
        cell.font      = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(xlsx_path)
    print("Spreadsheet updated: " + xlsx_path +
          "  (Results Table row " + str(target_row) +
          ", NLP row " + str(nlp_row) + ")")


# ── Spreadsheet creation helpers ──────────────────────────────────────────────

def _create_blank_spreadsheet(xlsx_path):
    # Builds a brand-new Results_Table.xlsx with both sheets correctly structured.

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()

    # Remove the default empty sheet
    wb.remove(wb.active)

    _add_results_sheet(wb)
    _add_nlp_sheet(wb)

    wb.save(xlsx_path)
    print("Created new spreadsheet: " + xlsx_path)


def _add_results_sheet(wb):
    # Adds the Results Table sheet with headers and placeholder rows.

    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet("Results Table")

    # Title
    ws.merge_cells("D1:L1")
    ws["D1"] = ("Proposed Results Table - Perceptron-Driven Anomaly Detection "
                "for Video Interviews")
    ws["D1"].font      = Font(name="Arial", size=12, bold=True)
    ws["D1"].alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[1].height = 40

    # Description rows
    ws["B3"] = "x1 = Eye-Gaze Score  (derived from YOLO-Pose: iris ratio, gaze stability)"
    ws["B4"] = "x2 = NLP Transcript Anomaly Score  (filler rate, repetition, length variation)"
    ws["B5"] = ("Perceptron: y = step(w1*x1 + w2*x2)  |  "
                "Fusion weights selected dynamically by rule engine")
    ws["B6"] = "Prompt Triggered: activated when output=1 AND confidence is low"
    ws["B7"] = "Final Classification: updated after prompt response is re-analysed"

    for row in range(3, 8):
        ws.cell(row=row, column=2).font = Font(name="Arial", size=9, italic=True)

    # Column headers  (row 4 = header row for data table)
    headers = [
        (4,  "Case ID"),
        (5,  "Eye-Gaze Score (x1)"),
        (6,  "Transcript Anomaly Score (x2)"),
        (7,  "Rules Engine Fusion Weights (w1, w2)"),
        (8,  "Weighted Sum\n(w1*x1 + w2*x2)"),
        (9,  "Perceptron Output"),
        (10, "System Decision"),
        (11, "Prompt Triggered"),
        (12, "Final Classification\n(Genuine / Anomaly Found / Genuine after prompt / Anomaly after prompt)"),
    ]

    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    for col, text in headers:
        cell = ws.cell(row=4, column=col)
        cell.value     = text
        cell.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[4].height = 42

    # Placeholder rows 5-9 (cases 1-5) with default weights
    default_weights = {5: "(0.6, 0.4)", 6: "(0.6, 0.4)", 7: "(0.6, 0.4)",
                       8: "(0.4, 0.6)", 9: "(0.5, 0.5)"}
    placeholder_fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    for row in range(5, 10):
        ws.cell(row=row, column=4).value = row - 4   # Case ID
        ws.cell(row=row, column=7).value = default_weights[row]
        for col in range(4, 13):
            cell = ws.cell(row=row, column=col)
            cell.fill      = placeholder_fill
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
        ws.row_dimensions[row].height = 20

    # Column widths
    col_widths = {4: 10, 5: 18, 6: 22, 7: 22, 8: 18, 9: 16, 10: 16, 11: 16, 12: 30}
    for col, width in col_widths.items():
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = width


def _add_nlp_sheet(wb):
    # Adds the NLP sheet with headers and placeholder rows.

    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet("NLP")

    # Title
    ws["B3"] = "NLP Analysis Results"
    ws["B3"].font = Font(name="Arial", size=12, bold=True)

    # Variable legend
    ws["B6"]  = "Filler"
    ws["C6"]  = "w1"
    ws["B7"]  = "Repetition"
    ws["C7"]  = "w2"
    ws["B8"]  = "Variation"
    ws["C8"]  = "w3"
    ws["B11"] = "Weightings Range"
    for val, row in [(0.25, 12), (0.5, 13), (0.75, 14), (1, 15), (1.25, 16), (1.5, 17)]:
        ws.cell(row=row, column=2).value = val

    # Column headers for test results (row 44)
    nlp_headers = [
        (5,  "Test"),
        (6,  "Filler Rate"),
        (7,  "Acceptable Band\n(Filler)"),
        (8,  "Repetition"),
        (9,  "Acceptable Band\n(Repetition)"),
        (10, "Variation"),
        (11, "Acceptable Band\n(Variation)"),
        (12, "Rules Engine NLP\nWeights (w1, w2, w3)"),
        (13, "Weighted Sum (wx)"),
    ]
    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    for col, text in nlp_headers:
        cell = ws.cell(row=44, column=col)
        cell.value     = text
        cell.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[44].height = 36

    # NLP weight sets for cases 1-5
    nlp_weight_sets = {
        45: "(0.4, 0.4, 0.2)",
        46: "(0.5, 0.3, 0.2)",
        47: "(0.3, 0.5, 0.2)",
        48: "(0.5, 0.4, 0.1)",
        49: "(0.4, 0.5, 0.1)",
    }
    placeholder_fill = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
    for row in range(45, 50):
        ws.cell(row=row, column=5).value  = row - 44
        ws.cell(row=row, column=12).value = nlp_weight_sets[row]
        for col in range(5, 14):
            cell = ws.cell(row=row, column=col)
            cell.fill      = placeholder_fill
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 18

    # Column widths
    col_widths = {5: 8, 6: 14, 7: 18, 8: 14, 9: 18, 10: 14, 11: 18, 12: 24, 13: 16}
    for col, width in col_widths.items():
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = width


# ── Helpers ───────────────────────────────────────────────────────────────────

def _band_label(value, low, high):
    if value <= low:
        return "Acceptable"
    elif value <= high:
        return "Borderline"
    else:
        return "Elevated"
